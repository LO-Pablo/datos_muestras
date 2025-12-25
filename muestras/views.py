from django.http import HttpResponse, FileResponse
from .models import Muestra, Localizacion, Estudio, Envio, Documento, historial_estudios, historial_localizaciones,agenda_envio, registro_destruido, Congelador, Estante, Rack,Caja, Subposicion
from django.template import loader
from .forms import MuestraForm, LocalizacionForm, UploadExcel, archivar_muestra_form, DocumentoForm, EstudioForm, Centroform, Congeladorform
from django.db import transaction
from django.contrib import messages  
from django.shortcuts import render,redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
import pandas as pd
import io,base64
from reportlab.pdfgen import canvas
from django.conf import settings
import openpyxl,os
from django.db.models import Q
from django.db import IntegrityError, ProgrammingError
from django.utils import timezone 
from django.contrib.auth.models import User
from numbers import Number
from collections import defaultdict
from django.db.models import Count, Q, Prefetch
@login_required
def principal(request):
    # Vista principal de la aplicación, muestra una página de bienvenida
    template = loader.get_template('principal.html')
    return HttpResponse(template.render(request=request))
@login_required

# Vistas para Muestras
@permission_required('muestras.can_view_muestras_web')
def muestras_todas(request):
    # Vista que muestra todas las muestras, requiere que el usuario esté autenticado
    muestras = Muestra.objects.prefetch_related('localizacion')
    # Filtrado de muestras si se proporcionan parámetros de búsqueda
    field_names = [f.name for f in Muestra._meta.local_fields if f.name not in ('id','estudio','estado_actual')]
    fields_loc = [f.name for f in Localizacion._meta.local_fields if f.name not in ('id','muestra')]
    field_names_readable = ['Id del individuo','Nombre dado por el laboratorio','Material','Volumen actual','Unidad de volumen','Concentración actual','Unidad de concentración','Masa actual','Unidad de masa','Fecha de extracción','Fecha de llegada','Observaciones','Estado inicial','Centro de procedencia','Lugar de procedencia']
    field_names_readable_dict = {k:v for (k,v) in zip(field_names,field_names_readable)}
    if request.user.groups.filter(name='Investigadores'):
        muestras = Muestra.objects.filter(Q(estudio__investigadores_asociados__username=request.user))
    for field in field_names:
        if request.GET.get(field):
            filter_kwargs = {f"{field}__icontains": request.GET[field]}
            muestras = muestras.filter(**filter_kwargs)
    if request.GET.get('estudio'):
        filtro_estudio = request.GET['estudio']
        muestras = muestras.filter(estudio__nombre_estudio__icontains=filtro_estudio)

    '''
    # Crear un PDF con las muestras filtradas
    if request.GET.get('crear_pdf'):    
        buffer = BytesIO()
        p = canvas.Canvas(buffer)
        y = 800
        p.setFont("Helvetica", 16)
        p.drawString(30,y, "Listado de Muestras")
        p.setFont("Helvetica", 12)
        y -= 30
        p.drawString(30, y, "ID Individuo")
        p.drawString(150, y, "Nombre Laboratorio")
        p.drawString(300, y, "Localización")
        y-= 30
        for muestra in muestras:
            p.drawString(30, y, muestra.id_individuo)
            p.drawString(150, y, muestra.nom_lab)
            p.drawString(300, y, str(muestra.localizacion.first()) if muestra.localizacion.exists() else 'No archivada')
            y -= 20
            if y < 50:
                p.showPage()
                y = 800
        p.save()
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='listado_muestras.pdf')
    '''
     # Crear un Excel con las muestras filtradas 
    if request.GET.get('exportar_excel'):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="listado_muestras.xlsx"'
        wb = openpyxl.load_workbook(os.path.join(settings.BASE_DIR, 'datos_prueba', 'globalstaticfiles', 'listado_muestras.xlsx'))
        ws = wb.active
        row_num = 2
        for muestra in muestras:
            col_num = 1
            
            for field in field_names:
                value = muestra.__dict__[field]
                if value is None:
                    value = ''
                ws.cell(row_num, col_num).value= str(value)
                col_num += 1
            value = muestra.estudio.nombre_estudio if muestra.estudio else ''
            ws.cell(row_num, col_num).value= str(value)
            col_num += 1
            try:
                loc = Localizacion.objects.get(muestra=muestra.nom_lab)
                for field in fields_loc:
                    value = loc.__dict__[field]
                    if value is None:
                        value = ''
                    ws.cell(row_num, col_num).value= str(value)
                    col_num += 1
                row_num += 1
            except:
                row_num += 1
        wb.save(response)
        return response
    template = loader.get_template('muestras_todas.html')
    context = {    
        'muestras': muestras,
        'field_names': field_names,
        'field_names_readable_dict': field_names_readable_dict
    }
    return HttpResponse(template.render(context, request))
@login_required
@permission_required('muestras.can_view_muestras_web')
def acciones_post(request):
    if request.method=="POST":
        muestras_seleccionadas = request.POST.getlist('muestra_id')
        if 'estudio' in request.POST:
            if muestras_seleccionadas:
                request.session['muestras_estudio']=muestras_seleccionadas
                return redirect('seleccionar_estudio')
        elif 'eliminar' in request.POST:
            if muestras_seleccionadas:
                muestras_a_procesar = Muestra.objects.filter(id__in=muestras_seleccionadas)
                for muestra in muestras_a_procesar:
                    eliminar_muestra(request, muestra.id_individuo, muestra.nom_lab) 
        elif 'envio' in request.POST:
            if muestras_seleccionadas:
                if 'muestras_envio' in request.session:
                    del request.session['muestras_envio']
                for muestra in muestras_seleccionadas:
                    if Muestra.objects.get(id=muestra).estado_actual == 'Destruida':
                        muestras_seleccionadas.remove(muestra)
                request.session['muestras_envio']=muestras_seleccionadas
                return redirect('agenda')
        elif 'destruir' in request.POST:
            if muestras_seleccionadas:
                muestras_a_destruir = Muestra.objects.filter(id__in=muestras_seleccionadas)
                for sample in muestras_a_destruir:
                    sample.estado_actual = 'Destruida'
                    sample.volumen_actual = 0
                    sample.concentracion_actual = 0
                    sample.save()
                    if Localizacion.objects.filter(muestra=sample).exists():
                        loc = Localizacion.objects.get(muestra=sample)
                        loc.muestra = None
                        loc.save()
                    registro_destruccion = registro_destruido.objects.create(muestra = sample,
                                                                             fecha = timezone.now(),
                                                                             usuario = request.user)
                    registro_destruccion.save()
                    
    return redirect('muestras_todas')    
def detalles_muestra(request, nom_lab):
    # Vista que muestra los detalles de una muestra específica, requiere permiso para ver muestras
    muestra = Muestra.objects.get(nom_lab=nom_lab)
    template = loader.get_template('detalles_muestra.html')
    context = {
        'muestra': muestra,
    }
    return HttpResponse(template.render(context, request))
@permission_required('muestras.can_add_muestras_web')


def añadir_muestras(request):
    if request.method == 'POST':
        form_muestra = MuestraForm(request.POST)
        form_archivo = archivar_muestra_form(request.POST)
        if form_muestra.is_valid() and form_muestra.is_valid():
            muestra = form_muestra.save()
            try:
                form_archivo.instance.muestra=muestra
                form_archivo.save()
                messages.success(request, 'Muestra añadida correctamente')
            except ValueError:
                messages.error(request,'La localización donde se quiere archivar la muestra está ocupada por otra muestra, la muestra a archivar se guardará sin localización')
            return redirect('muestras_todas')
    else:
        form_muestra = MuestraForm()
        form_archivo = archivar_muestra_form()
    return render(request, 'añadir_muestras.html', {'form_muestra': form_muestra, 'form_archivo':form_archivo})


@permission_required('muestras.can_delete_muestras_web')
def eliminar_muestra(request, id_individuo, nom_lab):
    # Vista para eliminar una muestra, requiere permiso para eliminar muestras
    muestra = get_object_or_404(Muestra,id_individuo=id_individuo, nom_lab=nom_lab)
    muestra.delete()
    messages.success(request,'Muestras eliminadas correctamente')
    return redirect('muestras_todas')
@permission_required('muestras.can_add_muestras_web')
def upload_excel(request):
    if request.method=="POST":
        form = UploadExcel(request.POST, request.FILES)
        if 'confirmar' in request.POST:
            messages.success(request, 'Las muestras sin errores graves se han añadido correctamente')
            ids_to_delete = request.session.pop('ids_error_muestras', [])
            Muestra.objects.filter(nom_lab__in=ids_to_delete).delete()
            return redirect('muestras_todas')
        elif 'cancelar' in request.POST:
            # Eliminación de las muestras y localizaciones añadidas del excel
            ids_to_delete = request.session.pop('nuevos_ids', [])
            Muestra.objects.filter(id__in=ids_to_delete).delete()
            messages.error(request,'Las muestras no se han añadido')
            return redirect('muestras_todas')
        elif 'excel_file' in request.FILES:
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                excel_bytes = excel_file.read()
                request.session['excel_file_name'] = excel_file.name
                request.session['excel_file_base64']= base64.b64encode(excel_bytes).decode()
                excel_stream = io.BytesIO(excel_bytes)
                df = pd.read_excel(excel_stream)
                rename_columns = {
                    'ID Individuo': 'id_individuo', 
                    'Nombre Laboratorio': 'nom_lab',
                    'ID Material': 'id_material',
                    'Volumen Actual': 'volumen_actual',
                    'Unidad Volumen': 'unidad_volumen',
                    'Concentracion Actual': 'concentracion_actual',
                    'Unidad Concentracion': 'unidad_concentracion',
                    'Masa Actual': 'masa_actual',
                    'Unidad Masa': 'unidad_masa',
                    'Fecha Extraccion': 'fecha_extraccion',
                    'Fecha Llegada': 'fecha_llegada',
                    'Observaciones': 'observaciones',
                    'Estado Inicial': 'estado_inicial',
                    'Centro Procedencia': 'centro_procedencia',
                    'Lugar Procedencia': 'lugar_procedencia',
                    'Estado actual': 'estado_actual',
                    'Congelador': 'congelador', 
                    'Estante': 'estante',
                    'Posición del rack en el estante': 'posicion_rack_estante',
                    'Rack': 'rack',
                    'Posición de la caja en el rack': 'posicion_caja_rack',
                    'Caja': 'caja',
                    'Subposición': 'subposicion',
                    'Estudio':'estudio'
                    
                }
                df.rename(columns=rename_columns, inplace=True)
                errors = 0
                errors_loc = 0
                errores_estudio = 0
                campos_vacios = 0
                localizaciones_ocupadas = 0
                nuevos_ids = []
                nuevos_ids_loc = []
                ids_error_muestras = []
                ids_error_localizaciones = []
                ids_formato_incorrecto = []
                ids_campos_vacios =[]
                ids_localizaciones_ocupadas = []
                ids_muestras_duplicadas = []
                columna_errores_formato = []
                numero_registros = 0
                for _, row in df.iterrows():
                    def normalize_value(value):
                        if pd.isna(value) or value is None:
                             return None
                        return value
                    numero_registros += 1
                    fecha_extraccion=normalize_value(row['fecha_extraccion'])
                    fecha_llegada = normalize_value(row['fecha_llegada'])
                    volumen_actual = normalize_value(row['volumen_actual'])
                    concentracion_actual = normalize_value(row['concentracion_actual'])
                    masa_actual = normalize_value(row['masa_actual'])
                    for campo in ['volumen_actual', 'concentracion_actual', 'masa_actual']:
                        try:
                            float(row[campo])
                        except:
                            columna_errores_formato.append(df.columns.get_loc(campo))
                            ids_formato_incorrecto.append(row['nom_lab'])
                            errors += 1
                            if campo == 'volumen_actual':
                                volumen_actual = 1000
                            elif campo == 'concentracion_actual':
                                concentracion_actual = 1000
                            else: 
                                masa_actual = 1000
                    for campo in ['fecha_extraccion', 'fecha_llegada']:
                        try:
                            pd.to_datetime(row[campo], errors='raise')
                        except Exception:
                            columna_errores_formato.append(df.columns.get_loc(campo))
                            ids_formato_incorrecto.append(row['nom_lab'])
                            errors += 1
                            if campo == 'fecha_llegada':
                                fecha_llegada = '1000-01-01'
                            else: 
                                fecha_extraccion='1000-01-01'
                            

                    nombre_estudio_excel = row['estudio']
                    if pd.isna(nombre_estudio_excel):
                        estudio_instance = None
                    else:
                        try:
                            estudio_instance = Estudio.objects.get(nombre_estudio=nombre_estudio_excel)
                        except  Exception:
                            estudio_instance = None
                            errores_estudio += 1
                    if not Muestra.objects.filter(nom_lab=row['nom_lab']):
                        congelador = Congelador.objects.get(congelador = row['congelador'])
                        estante = Estante.objects.get(congelador=congelador, numero=row['estante'])
                        rack = Rack.objects.get(estante=estante, numero=row['rack'])
                        caja = Caja.objects.get(rack=rack, numero=row['caja'])
                        subposicion= Subposicion.objects.get(caja = caja, numero = str(row['subposicion']))
                        if subposicion.vacia == False:
                            errors_loc += 1
                            raise ValueError(
                                f'Subposición {subposicion.numero} ya está ocupada'
                            )
                           

                        muestra, created = Muestra.objects.update_or_create(
                            id_individuo=row['id_individuo'],
                            nom_lab=row['nom_lab'],
                            id_material=normalize_value(row['id_material']),
                            volumen_actual=volumen_actual,
                            unidad_volumen=normalize_value(row['unidad_volumen']),
                            concentracion_actual=concentracion_actual,
                            unidad_concentracion=normalize_value(row['unidad_concentracion']),
                            masa_actual=masa_actual,
                            unidad_masa=normalize_value(row['unidad_masa']),
                            fecha_extraccion=fecha_extraccion,
                            fecha_llegada = fecha_llegada,
                            observaciones= normalize_value(row['observaciones']),
                            estado_inicial=normalize_value(row['estado_inicial']),
                            centro_procedencia=normalize_value(row['centro_procedencia']),
                            lugar_procedencia=normalize_value(row['lugar_procedencia']),
                            estado_actual=normalize_value(row['estado_actual']),
                            estudio = estudio_instance,
                            )
                        subposicion.vacia = False
                        subposicion.muestra = muestra
                        subposicion.save()
                        if not created:
                            ids_error_muestras.append(muestra.nom_lab)
                            errors+=1
                        if estudio_instance != None:
                            historial_estudio = historial_estudios.objects.create(muestra=muestra, estudio=estudio_instance,
                                                                        fecha_asignacion=timezone.now(), usuario_asignacion=request.user)
                    
                            historial_estudio.save()
                        def normalize_charfield_value(value):
                            if pd.isna(value) or value is None:
                                return None
                            try:
                                if isinstance(value, (float, int)) and value == int(value):
                                    return str(int(value))

                            except ValueError:
                                return str(value).strip()
                        localizacion, loc_created = Localizacion.objects.update_or_create(
                            muestra = muestra,
                            congelador=normalize_charfield_value(row['congelador']),
                            estante=normalize_charfield_value(row['estante']), 
                            posicion_rack_estante=normalize_charfield_value(row['posicion_rack_estante']),
                            rack=normalize_charfield_value(row['rack']),
                            posicion_caja_rack=normalize_charfield_value(row['posicion_caja_rack']),
                            caja=normalize_charfield_value(row['caja']),
                            subposicion=normalize_value(row['subposicion'])    
                        )
                        if created:
                            nuevos_ids.append(muestra.id)
                        if loc_created:
                            nuevos_ids_loc.append(localizacion.id)
                            Localizacion.objects.filter(congelador = localizacion.congelador, 
                                                                estante = localizacion.estante,
                                                                posicion_rack_estante = localizacion.posicion_rack_estante,
                                                                rack = localizacion.rack,
                                                                posicion_caja_rack = localizacion.posicion_caja_rack,
                                                                caja = localizacion.caja,
                                                                subposicion = localizacion.subposicion,
                                                                muestra__isnull=True).delete()
                            historial_loc = historial_localizaciones.objects.create(muestra=muestra, localizacion=localizacion,
                                                                            fecha_asignacion=timezone.now(), usuario_asignacion=request.user)
                        
                        historial_loc.save()
                            
                        
                        redirect('upload_excel')
                    else:
                        ids_muestras_duplicadas.append(row['nom_lab'])
                        errors+=1

                    for column in df.columns:
                        if pd.isna(row[column]):
                            if column in [f.name for f in Muestra._meta.local_fields if f.name in ('nom_lab','id_individuo')] or column in [f.name for f in Localizacion._meta.local_fields]:
                                if column =='nom_lab':
                                    ids_error_muestras.append(None)
                                    errors+=1
                                    Muestra.objects.filter(nom_lab='nan').delete()
                                elif column =='id_individuo':
                                    ids_error_muestras.append(row["nom_lab"])
                                    errors += 1
                                    Muestra.objects.filter(id=nuevos_ids[len(nuevos_ids)-1]).delete()
                                else:
                                    ids_error_muestras.append(row["nom_lab"])
                                    errors_loc+=1
                                    if not pd.isna(row['nom_lab']):
                                        Muestra.objects.filter(id=nuevos_ids[len(nuevos_ids)-1]).delete()
                                        ## if not localizacion.id == None:
                                            ## localizacion.delete()
                            elif column in [f.name for f in Muestra._meta.local_fields if f.name not in ('nom_lab','id_individuo')]:
                                Muestra.objects.filter(nom_lab=row['nom_lab']).update(**{column : None})
                                ids_campos_vacios.append(row['nom_lab']) 
                                campos_vacios += 1
                    if Localizacion.objects.filter(
                        ~Q(muestra__nom_lab=row['nom_lab']) | Q(muestra__isnull=True),
                        congelador=row['congelador'],
                        estante=row['estante'], 
                        posicion_rack_estante=row['posicion_rack_estante'],
                        rack=row['rack'],
                        posicion_caja_rack=row['posicion_caja_rack'],
                        caja=row['caja'],
                        subposicion=normalize_value(row['subposicion'])
                    ):
                        ids_localizaciones_ocupadas.append(row['nom_lab'])
                        localizaciones_ocupadas += 1
                    if row['nom_lab'] in ids_formato_incorrecto:
                        muestra.delete()
                request.session['nuevos_ids'] = nuevos_ids
                request.session['nuevos_ids_loc'] = nuevos_ids_loc
                request.session['ids_error_muestras'] = ids_error_muestras
                request.session['ids_error_localizaciones'] = ids_error_localizaciones
                request.session['ids_formato_incorrecto'] = ids_formato_incorrecto
                request.session['ids_campos_vacios'] = ids_campos_vacios
                request.session['ids_localizaciones_ocupadas'] = ids_localizaciones_ocupadas
                request.session['columna_errores_formato'] = columna_errores_formato
                request.session['ids_muestras_duplicadas'] = ids_muestras_duplicadas
                messages.info(request, f'El excel subido tiene {numero_registros} registros.')
                if errors==0 and errors_loc==0:
                    messages.success(request, 'Y no tiene errores en ningún campo.')
                    if campos_vacios!=0:
                        messages.info(request,f"Aunque tiene {campos_vacios} campos vacios en algunas muestras.")
                    elif localizaciones_ocupadas!=0:
                        messages.info(request,f"Aunque hay {localizaciones_ocupadas} muestras que se están intentando archivar en una posición ocupada por otra muestra.")
                else:
                    messages.warning(request, f'Y contiene {errors} errores de los campos de las muestras, {errores_estudio} errores en el campo de estudios y {errors_loc} de los campos de localización.')
                return render(request, 'confirmacion_upload.html')
        elif 'excel_errores' in request.POST:
                    ids_error_muestras = request.session.get('ids_error_muestras', [])
                    ids_error_localizaciones = request.session.get('ids_error_localizaciones', [])
                    ids_formato_incorrecto = request.session.get('ids_formato_incorrecto', [])
                    ids_campos_vacios=request.session.get('ids_campos_vacios', [])
                    ids_localizaciones_ocupadas=request.session.get('ids_localizaciones_ocupadas', [])
                    ids_muestras_duplicadas = request.session.get('ids_muestras_duplicadas', [])
                    columna_errores_formato = request.session.get('columna_errores_formato',[])
                    excel_bytes = base64.b64decode(request.session.get('excel_file_base64'))
                    excel_file = io.BytesIO(excel_bytes)
                    wb = openpyxl.load_workbook(excel_file)
                    ws = wb.active
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        
                        if row[1].value in ids_error_muestras or row[1].value in ids_muestras_duplicadas:
                            for cell in row:
                                cell.fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
                        elif row[1].value in ids_formato_incorrecto:
                            for cell in row:
                                if cell.col_idx - 1 in columna_errores_formato:
                                    cell.fill = openpyxl.styles.PatternFill(start_color="FF8000", end_color="FF8000", fill_type = "solid")
                        elif row[1].value in ids_campos_vacios:
                            for cell in row:
                                if cell.value == None:
                                    cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
                        elif row[1].value in ids_localizaciones_ocupadas:
                            for cell in row:
                                cell.fill = openpyxl.styles.PatternFill(start_color="51D1F6", end_color="51D1F6", fill_type = "solid")
                       
                    output = io.BytesIO()    
                    wb.save(output)
                    wb.close()
                    response = HttpResponse(output.getvalue(),content_type='application/ms-excel')
                    response['Content-Disposition'] = 'attachment; filename="listado_errores.xlsx"'
                    return response        
 
    else:
        form = UploadExcel(request)     
    return render(request, 'upload_excel.html', {'form': form}) 
   
@permission_required('muestras.can_change_muestras_web')
def editar_muestra(request, id_individuo, nom_lab):
    # Vista para editar una muestra existente, requiere permiso para cambiar muestras
    muestra = Muestra.objects.get(id_individuo=id_individuo, nom_lab=nom_lab)
    if request.method == 'POST':
        form = MuestraForm(request.POST, instance=muestra)
        if form.is_valid():
            form.save()
            return redirect('detalles_muestra', id_individuo=form.instance.id_individuo, nom_lab=form.instance.nom_lab)
    else:
        form = MuestraForm(instance=muestra)
    return render(request, 'editar_muestra.html', {'form': form, 'muestra': muestra})

def descargar_plantilla(request,macro:int):
    # Vista para descargar la plantilla de Excel para subir localizaciones o muestras
    if macro == 0:
        plantilla_path = os.path.join(settings.BASE_DIR, 'datos_prueba', 'globalstaticfiles', 'plantilla_localizaciones.xlsx')
        if os.path.exists(plantilla_path):
            return FileResponse(open(plantilla_path, 'rb'), as_attachment=True, filename='plantilla_localizaciones.xlsx')
    elif macro == 1:
        plantilla_path = os.path.join(settings.BASE_DIR, 'datos_prueba', 'globalstaticfiles', 'plantilla_localizaciones_macros.xlsm')
        if os.path.exists(plantilla_path):
            return FileResponse(open(plantilla_path, 'rb'), as_attachment=True, filename='plantilla_localizaciones_macros.xlsm')
    elif macro == 2:
        plantilla_path = os.path.join(settings.BASE_DIR, 'datos_prueba', 'globalstaticfiles', 'plantilla_muestras.xlsx')
        if os.path.exists(plantilla_path):
            return FileResponse(open(plantilla_path, 'rb'), as_attachment=True, filename='plantilla_muestras.xlsx')
    elif macro == 3:
        plantilla_path = os.path.join(settings.BASE_DIR, 'datos_prueba', 'globalstaticfiles', 'plantilla_estudios.xlsx')
        if os.path.exists(plantilla_path):
            return FileResponse(open(plantilla_path, 'rb'), as_attachment=True, filename='plantilla_estudios.xlsx')
    else:
        return HttpResponse("La plantilla no se encuentra disponible.", status=404)
    
# Vistas para Localizacion
@login_required
@permission_required('muestras.can_view_localizaciones_web')
def localizaciones(request):
    # Vista que muestra todas las localizaciones, tengan o no muestra
    cajas_qs = Caja.objects.annotate(
        numero_muestras=Count(
            'subposiciones',
            filter=Q(subposiciones__vacia=False)
        )
    )


    congeladores = Congelador.objects.prefetch_related(
        Prefetch(
            'estantes__racks__cajas',
            queryset=cajas_qs
        ),
        'estantes__racks__cajas__subposiciones')


    

    template = loader.get_template('localizaciones_todas.html')

    context = {
        'congeladores':congeladores,
    }
    return HttpResponse(template.render(context, request))
@permission_required('muestras.can_add_localizaciones_web')
def upload_excel_localizaciones(request):
    if request.method=="POST":
        form = UploadExcel(request.POST, request.FILES)
        if 'confirmar' in request.POST:
            messages.success(request, 'Las localizaciones se han añadido correctamente.')
            
        elif 'cancelar' in request.POST:
            ids_to_delete = request.session.pop('nuevos_ids', [])
            Congelador.objects.filter(id__in=ids_to_delete).delete()
        elif 'excel_file' in request.FILES:
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                df = pd.read_excel(excel_file)
                rename_columns = {
                    'Congelador': 'congelador', 
                    'Estante': 'estante',
                    'Posición del rack en el estante': 'posicion_rack_estante',
                    'Rack': 'rack',
                    'Posición de la caja en el rack': 'posicion_caja_rack',
                    'Caja': 'caja',
                    'Subposición': 'subposicion'
                }
                df.rename(columns=rename_columns, inplace=True)
                errors = 0
                nuevos_ids = []
                for _, row in df.iterrows():
                  
                        congelador, _ = Congelador.objects.get_or_create(congelador = str(int(row['congelador'])))
                        nuevos_ids.append(congelador.id)
                        estante, _ = Estante.objects.get_or_create(congelador=congelador, numero=str(int(row['estante'])))
                        rack , _= Rack.objects.get_or_create(estante=estante, numero=str(int(row['rack'])), defaults = {'posicion_rack_estante':row['posicion_rack_estante']})
                        caja, _ = Caja.objects.get_or_create(rack=rack, numero=str(int(row['caja'])), defaults = {'posicion_caja_rack':row['posicion_caja_rack']})
                        subposicion, created = Subposicion.objects.get_or_create(caja = caja, numero = str(int(row['subposicion'])))
                        if not created:
                            messages.warning(
                                request,
                                f'La subposición {row["subposicion"]} ya existe en la caja {caja.numero}'
                            )
                            errors += 1
                            
                    
                request.session['nuevos_ids'] = nuevos_ids
                nuevos_ids = []
                if errors==0:
                    messages.success(request, 'El archivo excel es correcto.')
                else:
                    messages.warning(request, f'El archivo excel contiene {errors} errores.') 
                return redirect('localizaciones_todas') 
    else:
        form = UploadExcel(request)     
    return render(request, 'localizacion_nueva.html', {'form': form}) 

def detalles_congelador(request, nombre_congelador):
    freezer= Congelador.objects.filter(congelador=nombre_congelador)
    template=loader.get_template('detalles_congelador.html')
    return HttpResponse(template.render({'congelador':freezer[0]},request))
@permission_required('muestras.can_add_localizaciones_web')
def editar_congelador(request,nombre_congelador):
    congelador = Congelador.objects.filter(congelador=nombre_congelador)
    congelador=congelador[0]
    if request.method == 'POST':
        form = Congeladorform(request.POST, request.FILES, instance=congelador)
        if form.is_valid():
            form.save()
            return redirect('detalles_congelador', nombre_congelador = form.instance.congelador)
    else:
        form = Congeladorform(instance=congelador)
    return render(request, 'editar_congelador.html', {'form': form, 'congelador': congelador})

@permission_required('muestras.can_delete_localizaciones_web')
def eliminar_localizacion(request):
    if len(request.POST.getlist('congelador')) > 0:
        congelador_lista = request.POST.getlist('congelador')
        for i in range(len(congelador_lista)):
            Congelador.objects.get(id=congelador_lista[i]).delete()

    if len(request.POST.getlist('estante')) > 0:
        estante_lista = request.POST.getlist('estante')
        for i in range(len(estante_lista)):
            Estante.objects.get(id=estante_lista[i]).delete()
                
    if len(request.POST.getlist('rack')) > 0:
        rack_lista = request.POST.getlist('rack')
        for i in range(len(rack_lista)):
            Rack.objects.get(id=rack_lista[i]).delete()

    if len(request.POST.getlist('caja')) > 0:
        caja_lista = request.POST.getlist('caja')
        for i in range(len(caja_lista)):
            Caja.objects.get(id=caja_lista[i]).delete()

    if len(request.POST.getlist('subposicion')) > 0:
        subposicion_lista = request.POST.getlist('subposicion')
        for i in range(len(subposicion_lista)):
            Subposicion.objects.get(id=subposicion_lista[i]).delete()

    else:
        return redirect('localizaciones_todas')
    
    return redirect('localizaciones_todas')


def historial_localizaciones_muestra(request,muestra_id):
    muestra = Muestra.objects.get(id=muestra_id)
    historiales = historial_localizaciones.objects.filter(muestra=muestra).order_by('-fecha_asignacion')
    if muestra.estado_actual=='Destruida':
        estado_destruccion = registro_destruido.objects.get(muestra=muestra)
    else:
        estado_destruccion = None
    template = loader.get_template('historial_localizaciones.html')
    return HttpResponse(template.render({'historiales':historiales, 'muestra':muestra, 'estado_destruccion':estado_destruccion},request))

# Vistas relacionadas con el modelo estudio
@login_required
@permission_required('muestras.can_view_estudios_web')
def estudios_todos(request):
    if request.user.groups.filter(name='Investigadores'):
        estudios = Estudio.objects.filter(investigadores_asociados=request.user)
    else:
        estudios = Estudio.objects.all()
    template = loader.get_template('estudios_todos.html')
    context = {
        'estudios':estudios
    }
    return HttpResponse(template.render(context,request))
@permission_required('muestras.can_add_estudios_web')
def nuevo_estudio(request):
    if request.method == 'POST':
        form = EstudioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request,'Estudio añadido correctamente')
            return redirect('estudios_todos')
    else:
        form = EstudioForm()
    template = loader.get_template('nuevo_estudio.html')
    return HttpResponse(template.render({'form':form},request))
@permission_required('muestras.can_add_estudios_web')
def excel_estudios(request):
    if request.method=="POST":
        form = UploadExcel(request.POST, request.FILES)
        if 'confirmar' in request.POST:
            messages.success(request, 'Las estudios se han añadido correctamente')
            return redirect('estudios_todos')
        elif 'cancelar' in request.POST:
            ids_to_delete = request.session.pop('nuevos_ids', [])
            Estudio.objects.filter(id__in=ids_to_delete).delete()
            messages.error(request,'Los estudios no se han añadido')
            return redirect('estudios_todos')
        elif 'excel_file' in request.FILES:
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                excel_bytes = excel_file.read()
                request.session['excel_file_name'] = excel_file.name
                request.session['excel_file_base64']= base64.b64encode(excel_bytes).decode()
                excel_stream = io.BytesIO(excel_bytes)
                df = pd.read_excel(excel_stream)
                rename_columns = {
                    'Referencia del estudio': 'referencia_estudio', 
                    'Nombre del estudio': 'nombre_estudio',
                    'Descripción': 'descripcion_estudio',
                    'Fecha de inicio': 'fecha_inicio_estudio',
                    'Fecha de fin': 'fecha_fin_estudio',
                    'Investigador principal': 'investigador_principal',
                }
                df.rename(columns=rename_columns, inplace=True)
                nuevos_ids = []
                estudios_existentes = []
                fechas_mal = []
                errors = 0
                def normalize_value(value):
                    if pd.isna(value) or value is None:
                        return None
                    return value
                def convertir_fecha(value):
                        if pd.isna(value) or value is None:
                            return None
                        if isinstance(value,Number):
                            raise ValueError
                        fecha = pd.to_datetime(value, errors='ignore')
                        return fecha.date()
                for _, row in df.iterrows():
                    try:
                        fecha_inicio_estudio = convertir_fecha(row['fecha_inicio_estudio'])
                        fecha_fin_estudio = convertir_fecha(row['fecha_fin_estudio'])
                        investigador_principal = normalize_value(row['investigador_principal'])
                        descripcion_estudio = normalize_value(row['descripcion_estudio'])
                        referencia_estudio = normalize_value(row['referencia_estudio'])
                    
                        estudio,created  = Estudio.objects.update_or_create(
                            nombre_estudio=row['nombre_estudio'],
                            defaults={
                                'referencia_estudio': referencia_estudio,
                                'descripcion_estudio': descripcion_estudio,
                                'fecha_inicio_estudio': fecha_inicio_estudio,
                                'fecha_fin_estudio': fecha_fin_estudio,
                                'investigador_principal': investigador_principal
                            }
                        )

                        if created:
                                nuevos_ids.append(estudio.id)
                        else:
                            messages.info(request, f'El estudio {estudio} ya existe, el excel no se ha procesado correctamente')
                            errors+=1
                            estudios_existentes.append(estudio.nombre_estudio)
                    except (ValueError, TypeError):
                        errors += 1
                        fechas_mal.append(row['nombre_estudio'])
                        continue
                   
                request.session['nuevos_ids'] = nuevos_ids
                request.session['estudios_existentes'] = estudios_existentes
                request.session['fechas_mal'] = fechas_mal
                if errors==0:
                    messages.success(request, 'El archivo excel es correcto.')
                else:
                    messages.warning(request, f'El archivo excel contiene {errors} errores.') 
                return render(request, 'confirmacion_upload_estudios.html')
        elif 'excel_errores' in request.POST:
                    estudios_existentes = request.session.get('estudios_existentes', [])
                    fechas_mal = request.session.get('fechas_mal', [])
                    excel_bytes = base64.b64decode(request.session.get('excel_file_base64'))
                    excel_file = io.BytesIO(excel_bytes)
                    wb = openpyxl.load_workbook(excel_file)
                    ws = wb.active
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        if row[1].value in estudios_existentes:
                            for cell in row:
                                cell.fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
                        elif row[1].value in fechas_mal:
                            for cell in row:
                                cell.fill = openpyxl.styles.PatternFill(start_color="FF8000", end_color="FF8000", fill_type = "solid")
                    output = io.BytesIO()    
                    wb.save(output)
                    wb.close()
                    response = HttpResponse(output.getvalue(),content_type='application/ms-excel')
                    response['Content-Disposition'] = 'attachment; filename="listado_errores.xlsx"'
                    return response      
    else: 
        form= UploadExcel()
    return render(request, 'upload_excel_estudios.html', {'form': form}) 
@permission_required('muestras.can_change_estudios_web')
def editar_estudio(request, id_estudio):
    estudio = Estudio.objects.get(id=id_estudio)
    if request.method == 'POST':
        form = EstudioForm(request.POST, instance=estudio)
        if form.is_valid():
            form.save()
            messages.info(request,'El estudio se ha modificado correctamente')
            return redirect('estudios_todos')
    else:
        form = EstudioForm(instance=estudio)
    return render(request, 'editar_estudio.html', {'form': form, 'estudio': estudio})
@permission_required('muestras.can_delete_estudios_web')
def eliminar_estudio(request, id_estudio):
    estudio = get_object_or_404(Estudio,id=id_estudio)
    estudio.delete()
    messages.success(request,'Estudio eliminado correctamente')
    return redirect('estudios_todos')
@permission_required('muestras.can_change_estudios_web')
def seleccionar_estudio(request):
    estudios = Estudio.objects.all()
    template = loader.get_template('seleccionar_estudio.html')
    return HttpResponse(template.render({'estudios':estudios},request))
@permission_required('muestras.can_change_estudios_web')
def añadir_muestras_estudio(request):
    if request.method == 'POST':
        muestras = request.session.get('muestras_estudio', [])
        muestras=Muestra.objects.filter(id__in=muestras)
        if len(request.POST.getlist('desasociar')) ==1:
            for muestra in muestras:
                if muestra.estado_actual != 'Destruida':
                    muestra.estudio = None
                    muestra.save()
                    historial = historial_estudios.objects.create(
                            muestra = muestra,
                            estudio = None,
                            fecha_asignacion = timezone.now(),
                            usuario_asignacion = request.user
                        )
                    historial.save()

            return redirect('muestras_todas')
        ids_estudios = request.POST.getlist('estudio_nombre')
        for study in ids_estudios:
            studio = Estudio.objects.get(nombre_estudio=study)
            for muestra in muestras:
                if muestra.estado_actual != 'Destruida':
                    muestra.estudio = studio
                    muestra.save()
                    if historial_estudios.objects.filter(muestra=muestra,estudio=studio).exists():
                        pass
                    else:   
                        historial = historial_estudios.objects.create(
                            muestra = muestra,
                            estudio = studio,
                            fecha_asignacion = timezone.now(),
                            usuario_asignacion = request.user
                        )
                        historial.save()
        if 'muestras_estudio' in request.session:
            del request.session['muestras_estudio']
        return redirect('muestras_todas')
    return redirect('muestras_todas')

def historial_estudios_muestra(request,muestra_id):
    muestra = Muestra.objects.get(id=muestra_id)
    historiales = historial_estudios.objects.filter(muestra=muestra).order_by('-fecha_asignacion')
    template = loader.get_template('historial_estudios.html')
    return HttpResponse(template.render({'historiales':historiales, 'muestra':muestra},request))
@permission_required('muestras.can_view_estudios_web')
def repositorio_estudio(request, id_estudio):
    estudio = Estudio.objects.get(id=id_estudio)
    documentos = Documento.objects.filter(estudio = estudio, eliminado= False)
    request.session['id'] = id_estudio
    # Filtrado opcional por usuario
    usuario = request.GET.get('usuario')
    if usuario:
        documentos = documentos.filter(usuario_subida__username=usuario)
    # Filtrado opcional por categoría
    categoria = request.GET.get('categoria')
    if categoria:
        documentos = documentos.filter(categoria__icontains=categoria)
    for doc in documentos:    
        if request.GET.get(f'{doc.id}'):
            eliminar_documento(request, doc.id)
    template = loader.get_template('repositorio_estudio.html')
    return HttpResponse(template.render({'documentos':documentos, 'id':estudio.id},request))

def subir_documento(request, id_estudio):
    estudio = Estudio.objects.get(id = id_estudio)
    if request.method == 'POST':
        form = DocumentoForm(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.usuario_subida = request.user
            doc.estudio = estudio
            doc.save()
            return redirect('repositorio_estudio', id_estudio=estudio.id)
        else:
            messages.error(request, 'Hubo un error al subir el documento.')
    else:
        form = DocumentoForm()
    template = loader.get_template('subir_documento.html')
    return HttpResponse(template.render({'form':form, 'estudio':estudio},request))

def descargar_documento(request, documento_id,id):
    doc = Documento.objects.get(pk=documento_id, eliminado=False)      
    return FileResponse(open(doc.archivo.path, 'rb'), as_attachment=True, filename=os.path.basename(doc.archivo.name))

def eliminar_documento(request):
    ids_documento = request.POST.getlist('doc_id')
    for element in ids_documento:
        try:
            doc = Documento.objects.get(pk=element, eliminado=False)
            doc.eliminado = True
            doc.fecha_eliminacion = timezone.now()
            doc.save()
            return redirect('repositorio_estudio', id_estudio=doc.estudio.id)
        except:
            return redirect('repositorio_estudio', id_estudio=doc.estudio.id)
    return redirect('repositorio_estudio', id=request.session.get('id'))

# Vistas relacionadas con el envio de muestras
@permission_required('muestras.can_change_muestras_web')
def formulario_envios(request,centro):
    muestras_envio = request.session.get('muestras_envio', [])
    centro_envio = agenda_envio.objects.get(id=centro)
    muestras = Muestra.objects.filter(id__in=muestras_envio, volumen_actual__gt=0)
    template = loader.get_template('formulario_envios.html')
    return HttpResponse(template.render({'muestras':muestras,'centro':centro_envio},request))

def upload_excel_envios(request,centro):
    if request.method=='POST':
        form = UploadExcel(request.POST, request.FILES)
        if 'descargar_excel_envio' in request.POST:
            centro_envio = agenda_envio.objects.get(id=centro)
            muestras = request.session.get('muestras_envio',[])
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="listado_envio.xlsx"'
            wb = openpyxl.load_workbook(os.path.join(settings.BASE_DIR, 'datos_prueba', 'globalstaticfiles', 'plantilla_envios.xlsx'))
            ws = wb.active
            row_num = 2
            for muestra in muestras:
                sample = Muestra.objects.get(id=muestra)
                if sample.estado_actual != 'Destruida':
                    ws.cell(row_num,1).value=str(sample.nom_lab)
                    ws.cell(row_num,2).value=str(sample.volumen_actual) + ' ' + str(sample.unidad_volumen)
                    ws.cell(row_num,3).value=str(sample.concentracion_actual) + ' ' + str(sample.unidad_concentracion)
                    ws.cell(row_num,5).value=str(sample.unidad_volumen)
                    ws.cell(row_num,7).value=str(sample.unidad_concentracion)
                    ws.cell(row_num,8).value=str(centro_envio.centro)
                    ws.cell(row_num,9).value=str(centro_envio.lugar)
                    row_num +=1 
            wb.save(response)
            return response
        elif 'excel_file' in request.FILES:
            if form.is_valid():
                ids_errores_envio = []
                errores_envio= 0
                errores_campos_vacios=0
                errores_formato = 0
                ids_errores_formato = []
                ids_errores_campos_vacios = []
                numero_registros = 0
                excel_file = request.FILES['excel_file']
                excel_bytes = excel_file.read()
                request.session['excel_file_name'] = excel_file.name
                request.session['excel_file_base64']= base64.b64encode(excel_bytes).decode()
                excel_stream = io.BytesIO(excel_bytes)
                df = pd.read_excel(excel_stream)
                rename_columns = {
                    'Muestra':'muestra',
                    'Volumen enviado':'volumen_enviado', 
                    'Concentración enviada':'concentracion_enviada',
                    'Centro de destino':'centro_destino',
                    'Lugar de destino':'lugar_destino'
                }
                df.rename(columns=rename_columns, inplace=True)
                for _, row in df.iterrows():
                    try:
                        instancia_muestra = Muestra.objects.get(nom_lab=row['muestra'])
                        envio = Envio.objects.create(muestra=instancia_muestra,
                                                    volumen_enviado=row['volumen_enviado'],
                                                    unidad_volumen_enviado=instancia_muestra.unidad_volumen,
                                                    concentracion_enviada=row['concentracion_enviada'],
                                                    centro_destino=row['centro_destino'],
                                                    unidad_concentracion_enviada=instancia_muestra.unidad_concentracion,
                                                    lugar_destino=row['lugar_destino'],
                                                    fecha_envio=timezone.now(),
                                                    usuario_envio=request.user
                                                    )
                        envio.save()
                        numero_registros += 1
                        if float(row['volumen_enviado']) == instancia_muestra.volumen_actual:
                            instancia_muestra.volumen_actual = 0
                            instancia_muestra.concentracion_actual = 0
                            instancia_muestra.estado_actual = 'Enviada'
                            instancia_muestra.save()
                            if Localizacion.objects.filter(muestra=instancia_muestra).exists():
                                loc = Localizacion.objects.get(muestra=instancia_muestra)
                                loc.muestra = None
                                loc.save()
                        elif float(row['volumen_enviado']) > instancia_muestra.volumen_actual or float(row['volumen_enviado']) <= 0:
                            ids_errores_envio.append(instancia_muestra.nom_lab)
                            errores_envio += 1
                            envio.delete()
                        else:
                            instancia_muestra.volumen_actual -= float(row['volumen_enviado'])
                            instancia_muestra.estado_actual = 'Parcialmente enviada'
                            instancia_muestra.save()
                    except ValueError:
                        errores_formato += 1
                        ids_errores_formato.append(instancia_muestra.nom_lab)
                        numero_registros+=1
                    except ProgrammingError:
                        errores_campos_vacios += 1
                        ids_errores_campos_vacios.append(instancia_muestra.nom_lab)
                        numero_registros+=1
                if 'muestras_envio' in request.session:
                    del request.session['muestras_envio']
                request.session['ids_errores_formato'] = ids_errores_formato
                request.session['ids_errores_envio'] = ids_errores_envio
                request.session['ids_errores_campos_vacios']= ids_errores_campos_vacios
                messages.info(request, f'El excel subido tiene {numero_registros} registros.')
                if errores_envio==0 and errores_campos_vacios==0 and errores_formato==0:
                    messages.success(request, 'Y no tiene errores en ningún campo.')
                else:
                    messages.warning(request, f'Y contiene {errores_envio} errores en el volumen de envio de algunas muestras, {errores_formato} errores de formato y {errores_campos_vacios} campos vacios.')
                return render(request,'confirmacion_upload_envio.html') 
        elif 'excel_errores' in request.POST:
                ids_errores_envio = request.session.get('ids_errores_envio', [])
                ids_errores_campos_vacios= request.session.get('ids_errores_campos_vacios',[])
                ids_errores_formato=request.session.get('ids_errores_formato',[])
                excel_bytes = base64.b64decode(request.session.get('excel_file_base64'))
                excel_file = io.BytesIO(excel_bytes)
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):

                    if row[0].value in ids_errores_campos_vacios:
                        for cell in row:
                            cell.fill = openpyxl.styles.PatternFill(start_color="FF8000", end_color="FF8000", fill_type = "solid")
                    elif row[0].value in ids_errores_formato:
                        for cell in row:
                            cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
                    elif row[0].value in ids_errores_envio:
                        for cell in row:
                            cell.fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
                    
                output = io.BytesIO()    
                wb.save(output)
                wb.close()
                response = HttpResponse(output.getvalue(),content_type='application/ms-excel')
                response['Content-Disposition'] = 'attachment; filename="listado_errores_envio.xlsx"'
                return response         
    else:
        form = UploadExcel(request)
    template = loader.get_template('upload_excel_envios.html')     
    return HttpResponse(template.render({'form': form},request))

def registrar_envio(request,centro):
    if request.method=='POST':
        centro_envio = agenda_envio.objects.get(id=centro)
        muestras = request.session.get('muestras_envio', [])
        volumen_enviado_form = request.POST.getlist('volumen_enviado')
        concentracion_enviada_form = request.POST.getlist('concentracion_enviada')
        centro_destino_form = centro_envio.centro
        lugar_destino_form = centro_envio.lugar
        iterar = 0
        for muestra in muestras:
            instancia_muestra = Muestra.objects.get(id=muestra)
            envio = Envio.objects.create(
                muestra=instancia_muestra,
                fecha_envio=timezone.now(),
                volumen_enviado = volumen_enviado_form[iterar],
                unidad_volumen_enviado = instancia_muestra.unidad_volumen,
                concentracion_enviada = concentracion_enviada_form[iterar],
                unidad_concentracion_enviada = instancia_muestra.unidad_concentracion,
                centro_destino = centro_destino_form,
                lugar_destino=lugar_destino_form,
                usuario_envio = request.user
            )
            envio.save()
            if float(volumen_enviado_form[iterar]) >= instancia_muestra.volumen_actual:
                instancia_muestra.volumen_actual = 0
                instancia_muestra.concentracion_actual = 0
                instancia_muestra.estado_actual = 'Enviada'
                instancia_muestra.save()
                if Localizacion.objects.filter(muestra=instancia_muestra).exists():
                    loc = Localizacion.objects.get(muestra=instancia_muestra)
                    loc.muestra = None
                    loc.save()
            else:
                instancia_muestra.volumen_actual -= float(volumen_enviado_form[iterar])
                instancia_muestra.estado_actual = 'Parcialmente enviada'
                instancia_muestra.save()
            iterar += 1
        if 'muestras_envio' in request.session:
            del request.session['muestras_envio']
        return redirect('muestras_todas')
    return redirect('formulario_envios')

def historial_envios(request,muestra_id):
    sample = Muestra.objects.get(id=muestra_id)
    envios = Envio.objects.filter(muestra=sample).order_by('-fecha_envio')
    volumen_original = sample.volumen_actual + sum(envio.volumen_enviado for envio in envios)
    volumen_restante = sample.volumen_actual
    template = loader.get_template('historial_envios.html')
    context = {
        'muestra':sample,
        'envios':envios,
        'volumen_original':volumen_original,
        'volumen_restante':volumen_restante
    }
    return HttpResponse(template.render(context,request))

def agenda(request):
    agenda_envios = agenda_envio.objects.all()
    template = loader.get_template('agenda.html')
    return HttpResponse(template.render({'agenda':agenda_envios},request))

def nuevo_centro(request):
    if request.method == 'POST':
        form = Centroform(request.POST)
        if form.is_valid():
            form.save()
            return redirect('agenda')
        else:
            messages.error(request, 'Hubo un error al añadir el centro.')
    else:
        form = Centroform()
    template = loader.get_template('nuevo_centro.html')
    return HttpResponse(template.render({'form':form},request))

def editar_centro(request, id_centro):
    centro = agenda_envio.objects.get(id=id_centro)
    if request.method == 'POST':
        form = Centroform(request.POST, instance=centro)
        if form.is_valid():
            form.save()
            return redirect('agenda')
    else:
        form = Centroform(instance=centro)
    return render(request, 'editar_centro.html', {'form': form, 'centro': centro})

def eliminar_centro(request):
    if request.method=="POST":
        ids = request.POST.getlist('ids_centro')
        for centro_id in ids:
            centro = agenda_envio.objects.get(id=centro_id)
            centro.delete()
    return redirect('agenda')